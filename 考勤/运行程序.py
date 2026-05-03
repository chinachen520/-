
import re
import sys
import os
import io
from collections import defaultdict

if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

try:
    import openpyxl
except ImportError:
    print("缺少 openpyxl 库，请执行: pip install openpyxl")
    input("\n按回车退出...")
    sys.exit(1)


STU_PATH = os.path.join(os.path.dirname(__file__), "学号.xlsx")
KQ_PATH = os.path.join(os.path.dirname(__file__), "考勤表.xlsx")
TXT_PATH = os.path.join(os.path.dirname(__file__), "本周考勤.txt")


STU_SHEET = "Sheet1"
STU_NAME_COL = 5
STU_ID_COL = 4


KQ_SHEET = "Sheet1"
KQ_START_ROW = 4
KQ_CLASS_VAL = "325人工2"
KQ_COL_MAP = {
    "请假": 4,
    "早退": 7,
    "迟到": 6,
    "旷课": 5,
}

def parse_attendance(lines):

    stats = defaultdict(lambda: {"请假": 0, "早退": 0, "迟到": 0, "旷课": 0, "公假": 0})
    categories = ["请假", "早退", "迟到", "旷课", "公假"]

    for line in lines:
        line = line.strip()
        if not line:
            continue

        for cat in categories:

            pattern = re.compile(
                rf'{cat}:\d+\s*[（(]\s*([^）)]+?)\s*[）)]',
                re.UNICODE
            )
            m = pattern.search(line)
            if m:
                names_str = m.group(1)
                names = [n.strip() for n in re.split(r'[、，,]', names_str) if n.strip()]
                for name in names:
                    stats[name][cat] += 1

    return dict(stats)


def load_student_ids(xlsx_path):

    if not os.path.exists(xlsx_path):
        print(f"[!] 未找到学号表: {xlsx_path}")
        return {}

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[STU_SHEET]
    stu_map = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        name = row[STU_NAME_COL - 1] if len(row) >= STU_NAME_COL else None
        sid = row[STU_ID_COL - 1] if len(row) >= STU_ID_COL else None
        if name and sid:
            stu_map[str(name).strip()] = str(sid).strip()
    wb.close()
    return stu_map


def fill_attendance_table(stats, stu_map, kq_path, dry_run=False):

    if not os.path.exists(kq_path):
        print(f"[X] 未找到考勤表: {kq_path}")
        return 0

    wb = openpyxl.load_workbook(kq_path)
    ws = wb[KQ_SHEET]

    merged_cells = list(ws.merged_cells.ranges)

    for r in range(KQ_START_ROW, KQ_START_ROW + 200):
        has_data = False
        for c in range(1, 8):
            cell = ws.cell(row=r, column=c)

            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                continue
            if cell.value is not None:
                has_data = True
                break
        if not has_data:
            break
        for c in range(1, 8):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                continue
            cell.value = None


    sorted_names = sorted(
        stats.keys(),
        key=lambda n: (sum(stats[n].values()), n),
        reverse=True
    )

    row_idx = KQ_START_ROW
    unknown_names = []

    for name in sorted_names:
        sid = stu_map.get(name, "")
        if not sid:
            unknown_names.append(name)

        if dry_run:
            parts = [f"{k}={v}" for k, v in stats[name].items() if v > 0]
            print(f"  [{row_idx}] {name} ({sid}): {', '.join(parts)}")
        else:

            cell_b = ws.cell(row=row_idx, column=2)
            if isinstance(cell_b, openpyxl.cell.cell.MergedCell):

                for mc in merged_cells:
                    if mc.min_row <= row_idx <= mc.max_row and mc.min_col <= 2 <= mc.max_col:
                        ws.unmerge_cells(str(mc))
                        break

            cell_a = ws.cell(row=row_idx, column=1)
            if not isinstance(cell_a, openpyxl.cell.cell.MergedCell):
                ws.cell(row=row_idx, column=1).value = KQ_CLASS_VAL
            ws.cell(row=row_idx, column=2).value = name
            ws.cell(row=row_idx, column=3).value = sid
            for cat, col in KQ_COL_MAP.items():
                cnt = stats[name].get(cat, 0)
                if cnt > 0:
                    cell = ws.cell(row=row_idx, column=col)
                    if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                        cell.value = cnt
        row_idx += 1

    if not dry_run:
        wb.save(kq_path)
    wb.close()

    if unknown_names:
        print(f"\n[!] 以下 {len(unknown_names)} 人未在学号表中找到学号：")
        for n in unknown_names:
            print(f"    - {n}")

    return row_idx - KQ_START_ROW


def main():

    txt_path = TXT_PATH
    dry_run = False
    args = [a for a in sys.argv[1:] if a != sys.argv[0]]

    for a in args:
        if a == "--dry-run":
            dry_run = True
        elif not a.startswith("--"):
            txt_path = a

    if not os.path.exists(txt_path):
        print(f"[X] 未找到考勤文件: {txt_path}")
        print(f"    请先在 {TXT_PATH} 中填入考勤数据")
        input("\n按回车退出...")
        return


    with open(txt_path, "r", encoding="utf-8") as f:
        lines = f.readlines()

    print("=" * 60)
    print("  考勤登记工具")
    print("=" * 60)
    print(f"  考勤文件: {txt_path}")
    print(f"  学号表:   {STU_PATH}  ({'找到' if os.path.exists(STU_PATH) else '未找到'})")
    print(f"  考勤表:   {KQ_PATH}  ({'找到' if os.path.exists(KQ_PATH) else '未找到'})")
    print()


    stats = parse_attendance(lines)
    print(f"  解析到 {len(stats)} 人有考勤记录：")
    for name, cats in sorted(stats.items(), key=lambda x: -sum(x[1].values())):
        parts = [f"{k}={v}" for k, v in cats.items() if v > 0]
        print(f"    {name}: {', '.join(parts)}")

    print()

    if dry_run:
        print("  [预览模式 - 不写入表格]")
        print()
        fill_attendance_table(stats, {}, KQ_PATH, dry_run=True)
        print(f"\n[OK] 预览完成，共 {len(stats)} 人")
        return


    stu_map = load_student_ids(STU_PATH)


    count = fill_attendance_table(stats, stu_map, KQ_PATH)

    if count > 0:
        print(f"\n[OK] 完成！共填写 {count} 条记录到考勤表")
        print(f"     {KQ_PATH}")
    else:
        print("\n[!] 未写入任何数据")

    input("\n按回车退出...")


if __name__ == "__main__":
    main()
